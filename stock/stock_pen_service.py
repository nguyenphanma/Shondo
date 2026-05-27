"""
Service functions for stock_pen — called by Streamlit UI.
All heavy DB logic lives here; scripts (update_stock_pen.py) stay as standalone runners.
"""
import pandas as pd
from datetime import date
from sqlalchemy import text
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.queries import get_product_template

_MOVEMENTS_SQL = """
SELECT
    DATE(im.date)      AS success_date,
    imi.product_code   AS fdcode,
    imi.quantity       AS qty,
    CASE
        WHEN im.description LIKE '%KDC%'      THEN 'KDC'
        WHEN im.description LIKE '%KHO TỔNG%' THEN 'KDC'
        WHEN im.description LIKE '%KHO SỈ%'   THEN 'KDS'
        WHEN im.description LIKE '%ECOM%'     THEN 'ECOM'
        ELSE im.description
    END AS channel
FROM inventory_movement_items imi
LEFT JOIN inventory_movements im ON im.id = imi.movement_id
WHERE im.depot_id = 101011
  AND im.mode = 3
  AND im.date >= :from_date
  AND im.date <= :to_date
  AND (
        im.description LIKE '%KDC%'
        OR im.description LIKE '%KHO TỔNG%'
        OR im.description LIKE '%KHO SỈ%'
        OR im.description LIKE '%ECOM%'
      )
"""


# ── CHECK ─────────────────────────────────────────────────────────────────────

def check_new_movements(engine, from_date: str, to_date: str = None) -> dict:
    """
    So sánh movements trong khoảng với delivered hiện tại trong stock_pen.
    Returns dict: has_new_data, df_daily, df_compare, total_units, covered_months.
    """
    if to_date is None:
        to_date = str(date.today())

    with engine.connect() as conn:
        df_mv = pd.read_sql(
            text(_MOVEMENTS_SQL), conn,
            params={'from_date': from_date, 'to_date': to_date}
        )
        if df_mv.empty:
            return {
                'has_new_data': False, 'df_daily': pd.DataFrame(),
                'df_compare': pd.DataFrame(), 'total_units': 0, 'covered_months': []
            }

        df_mv['success_date'] = pd.to_datetime(df_mv['success_date'])
        df_mv['month'] = df_mv['success_date'].dt.month
        covered_months = sorted(df_mv['month'].unique().tolist())

        # Daily detail
        df_daily = (
            df_mv.groupby(['success_date', 'channel'])['qty'].sum()
            .reset_index()
            .rename(columns={'success_date': 'Ngày', 'channel': 'Kênh', 'qty': 'SL sản xuất'})
        )
        df_daily['Ngày'] = pd.to_datetime(df_daily['Ngày']).dt.strftime('%Y-%m-%d')

        # So sánh TỔNG (không per-channel vì cross-channel allocation làm lệch số theo kênh)
        del_cols = [f'delivered_{m}' for m in covered_months]
        sum_expr = ' + '.join(del_cols)
        row = conn.execute(
            text(f"SELECT SUM({sum_expr}) as total_stock FROM stock_pen")
        ).fetchone()
        total_stock = int(row[0] or 0)

        # Breakdown theo kênh chỉ để hiển thị tham khảo
        df_mv_by_channel = (
            df_mv.groupby('channel')['qty'].sum().reset_index()
            .rename(columns={'channel': 'Kênh', 'qty': 'SL xuất kho SX'})
        )

    total_mv = int(df_mv['qty'].sum())
    diff = total_mv - total_stock

    df_compare = pd.DataFrame([{
        'Tổng movements (xuất kho SX)': total_mv,
        'Tổng đã ghi vào stock_pen': total_stock,
        'Chênh lệch': diff,
    }])

    return {
        'has_new_data': diff > 0,
        'df_daily': df_daily,
        'df_mv_by_channel': df_mv_by_channel,
        'df_compare': df_compare,
        'total_units': total_mv,
        'covered_months': covered_months,
    }


# ── UPDATE ────────────────────────────────────────────────────────────────────

def run_update_stock_pen(engine, from_date: str, to_date: str = None) -> dict:
    """
    Full idempotent update: reset delivered for covered months → FIFO (same-channel
    then cross-channel) → insert NGOÀI ĐƠN → recalculate order_pen.
    Returns result summary dict.
    """
    if to_date is None:
        to_date = str(date.today())

    with engine.connect() as conn:
        df_mv = pd.read_sql(
            text(_MOVEMENTS_SQL), conn,
            params={'from_date': from_date, 'to_date': to_date}
        )
        df_stock_full = pd.read_sql(
            text("SELECT * FROM stock_pen WHERE channel != 'NGOÀI ĐƠN' ORDER BY channel, fdcode, year_ord, month_ord"),
            conn
        )

    if df_mv.empty:
        return {'success': False, 'message': 'Không có movements trong khoảng thời gian này.'}

    df_prod = (
        get_product_template(engine)[['fdcode', 'default_code', 'category', 'subcategory', 'size']]
        .drop_duplicates('fdcode')
    )

    df_mv['success_date'] = pd.to_datetime(df_mv['success_date'])
    df_mv['month'] = df_mv['success_date'].dt.month
    covered_months = sorted(df_mv['month'].unique().tolist())

    df_agg = df_mv.groupby(['fdcode', 'channel', 'month'], as_index=False)['qty'].sum()

    stock_keys = set(zip(df_stock_full['fdcode'], df_stock_full['channel']))
    mask = df_agg.apply(lambda r: (r['fdcode'], r['channel']) in stock_keys, axis=1)
    df_matched = df_agg[mask].copy()
    df_ngoai_base = (
        df_agg[~mask]
        .groupby(['fdcode', 'month'], as_index=False)['qty'].sum()
    )

    fixed_del_cols = ['delivered_old_year'] + [
        f'delivered_{m}' for m in range(1, 13) if m not in covered_months
    ]
    df_stock_full['_debt_before_reset'] = (
        df_stock_full['qty_ord'] - df_stock_full[fixed_del_cols].sum(axis=1)
    ).clip(lower=0)

    allocations: dict = {}
    remaining_debt_final: dict = {}
    channel_surplus: list = []

    for (fdc, ch), grp_mv in df_matched.groupby(['fdcode', 'channel']):
        orders = df_stock_full[
            (df_stock_full['fdcode'] == fdc) & (df_stock_full['channel'] == ch)
        ].copy().sort_values(['year_ord', 'month_ord'])
        if orders.empty:
            continue
        remaining_debt = orders['_debt_before_reset'].tolist()
        row_id_list = orders['id'].tolist()
        for _, mv_row in grp_mv.iterrows():
            delivery_month = int(mv_row['month'])
            remaining_qty = int(mv_row['qty'])
            for i, rid in enumerate(row_id_list):
                if remaining_qty <= 0:
                    break
                debt = remaining_debt[i]
                if debt <= 0:
                    continue
                to_alloc = min(remaining_qty, debt)
                allocations[(rid, delivery_month)] = allocations.get((rid, delivery_month), 0) + to_alloc
                remaining_debt[i] -= to_alloc
                remaining_qty -= to_alloc
            if remaining_qty > 0:
                channel_surplus.append({'fdcode': fdc, 'month': delivery_month, 'qty': remaining_qty})
        for rid, debt in zip(row_id_list, remaining_debt):
            remaining_debt_final[rid] = debt

    for _, row in df_stock_full.iterrows():
        if row['id'] not in remaining_debt_final:
            remaining_debt_final[int(row['id'])] = row['_debt_before_reset']

    cross_pool = channel_surplus.copy()
    for _, row in df_ngoai_base.iterrows():
        cross_pool.append({'fdcode': row['fdcode'], 'month': int(row['month']), 'qty': int(row['qty'])})

    fdcode_rows = (
        df_stock_full.sort_values(['year_ord', 'month_ord'])
        .groupby('fdcode')['id'].apply(list).to_dict()
    )

    true_ngoai: list = []
    if cross_pool:
        pool_df = pd.DataFrame(cross_pool).groupby(['fdcode', 'month'], as_index=False)['qty'].sum()
        for _, s_row in pool_df.iterrows():
            fdc = s_row['fdcode']
            delivery_month = int(s_row['month'])
            pool_qty = int(s_row['qty'])
            for rid in fdcode_rows.get(fdc, []):
                if pool_qty <= 0:
                    break
                debt = remaining_debt_final.get(rid, 0)
                if debt <= 0:
                    continue
                to_alloc = min(pool_qty, debt)
                allocations[(rid, delivery_month)] = allocations.get((rid, delivery_month), 0) + to_alloc
                remaining_debt_final[rid] -= to_alloc
                pool_qty -= to_alloc
            if pool_qty > 0:
                true_ngoai.append({'fdcode': fdc, 'month': delivery_month, 'qty': pool_qty})

    df_ngoai = (
        pd.DataFrame(true_ngoai, columns=['fdcode', 'month', 'qty'])
        if true_ngoai else pd.DataFrame(columns=['fdcode', 'month', 'qty'])
    )

    delivered_sum_sql = ' + '.join(['delivered_old_year'] + [f'delivered_{m}' for m in range(1, 13)])
    prod_lookup = df_prod.set_index('fdcode').to_dict('index')

    with engine.connect() as conn:
        for m in covered_months:
            conn.execute(text(f"UPDATE stock_pen SET delivered_{m} = 0 WHERE channel != 'NGOÀI ĐƠN'"))
        conn.execute(text("DELETE FROM stock_pen WHERE channel = 'NGOÀI ĐƠN'"))
        conn.commit()

        updated_rows = 0
        for (row_id, delivery_month), qty in allocations.items():
            col = f'delivered_{delivery_month}'
            r = conn.execute(
                text(f'UPDATE stock_pen SET {col} = {col} + :qty WHERE id = :id'),
                {'qty': qty, 'id': row_id}
            )
            updated_rows += r.rowcount

        ngoai_inserted = 0
        for fdc, grp in df_ngoai.groupby('fdcode'):
            all_cols = {f'delivered_{m}': 0 for m in range(1, 13)}
            for _, r in grp.iterrows():
                all_cols[f'delivered_{int(r["month"])}'] = int(r['qty'])
            meta = prod_lookup.get(fdc, {})
            cols_sql = ', '.join(all_cols.keys())
            vals_sql = ', '.join([f':{c}' for c in all_cols.keys()])
            conn.execute(
                text(f"""
                    INSERT INTO stock_pen
                        (channel, fdcode, category, subcategory, default_code, size,
                         month_ord, year_ord, qty_ord, qty_delivered_by_manu, order_pen,
                         delivered_old_year, {cols_sql})
                    VALUES
                        ('NGOÀI ĐƠN', :fdcode, :category, :subcategory, :default_code, :size,
                         0, 2026, 0, 0, 0, 0, {vals_sql})
                """),
                {**all_cols, 'fdcode': fdc,
                 'category': meta.get('category') or None,
                 'subcategory': meta.get('subcategory') or None,
                 'default_code': meta.get('default_code') or None,
                 'size': meta.get('size') or None}
            )
            ngoai_inserted += 1

        conn.execute(text(f"""
            UPDATE stock_pen SET
                qty_delivered_by_manu = ({delivered_sum_sql}),
                order_pen = GREATEST(0, qty_ord - ({delivered_sum_sql}))
            WHERE channel != 'NGOÀI ĐƠN'
        """))
        conn.commit()

    return {
        'success': True,
        'covered_months': covered_months,
        'updated_rows': updated_rows,
        'ngoai_inserted': ngoai_inserted,
        'total_mv': int(df_mv['qty'].sum()),
    }


# ── INSERT ORDERS ──────────────────────────────────────────────────────────────

def prepare_orders_df(uploaded_file) -> pd.DataFrame:
    """Parse uploaded Excel (sheet: DATA  ALL) into cleaned orders DataFrame."""
    df = pd.read_excel(uploaded_file, sheet_name='DATA  ALL')
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    cols = [
        'KÊNH BÁN', 'DANH MỤC', 'DANH MỤC CON', 'MÃ SP CHA', 'SIZE', 'Mã hàng',
        'ĐƠN ĐẶT HÀNG THÁNG', 'NĂM', 'SL ĐẶT', 'TỔNG TRẢ', 'SL CÒN NỢ\n(XƯỞNG)',
        'SL TRẢ\nNĂM 2023-2025',
    ] + [f'SL TRẢ T{str(i).zfill(2)}' for i in range(1, 13)]
    df = df[cols]

    numeric_cols = [
        'SL ĐẶT', 'TỔNG TRẢ', 'SL CÒN NỢ\n(XƯỞNG)', 'SL TRẢ\nNĂM 2023-2025',
    ] + [f'SL TRẢ T{str(i).zfill(2)}' for i in range(1, 13)]
    df[numeric_cols] = df[numeric_cols].fillna(0).astype(int)
    df[['MÃ SP CHA']] = df[['MÃ SP CHA']].apply(lambda x: x.str.upper())
    df['KÊNH BÁN'].replace({'CỬA HÀNG': 'KDC', 'BÁN SỈ': 'KDS'}, inplace=True)
    df['ĐƠN ĐẶT HÀNG THÁNG'].replace({'Nợ 2024': '12'}, inplace=True)
    df['SIZE'] = df['SIZE'].str.replace('Size ', '', regex=False)
    df['ĐƠN ĐẶT HÀNG THÁNG'] = (
        df['ĐƠN ĐẶT HÀNG THÁNG'].astype(str).str.extract(r'(\d{1,2})')[0].str.zfill(2)
    )

    df.rename(columns={
        'SL CÒN NỢ\n(XƯỞNG)': 'order_pen',
        'TỔNG TRẢ': 'qty_delivered_by_manu',
        'MÃ SP CHA': 'default_code',
        'KÊNH BÁN': 'channel',
        'DANH MỤC': 'category',
        'DANH MỤC CON': 'subcategory',
        'SIZE': 'size',
        'Mã hàng': 'fdcode',
        'ĐƠN ĐẶT HÀNG THÁNG': 'month_ord',
        'NĂM': 'year_ord',
        'SL ĐẶT': 'qty_ord',
        'SL TRẢ\nNĂM 2023-2025': 'delivered_old_year',
        **{f'SL TRẢ T{str(i).zfill(2)}': f'delivered_{i}' for i in range(1, 13)}
    }, inplace=True)

    return df


def insert_orders(engine, df: pd.DataFrame, replace_month: bool = False) -> dict:
    """
    Insert orders into stock_pen.

    replace_month=False (default): chỉ INSERT thêm — không xóa gì cả.
    replace_month=True: trước khi insert, xóa các dòng cũ có cùng
        (year_ord, month_ord, channel) với file upload — dùng khi đơn tháng đó bị sai.
        NGOÀI ĐƠN không bao giờ bị xóa.

    Returns dict: inserted, deleted.
    """
    deleted = 0
    if replace_month:
        combos = (
            df[['year_ord', 'month_ord', 'channel']]
            .drop_duplicates()
        )
        with engine.connect() as conn:
            for _, row in combos.iterrows():
                r = conn.execute(
                    text("""
                        DELETE FROM stock_pen
                        WHERE year_ord = :y AND month_ord = :m AND channel = :ch
                          AND channel != 'NGOÀI ĐƠN'
                    """),
                    {'y': int(row['year_ord']), 'm': int(row['month_ord']), 'ch': str(row['channel'])}
                )
                deleted += r.rowcount
            conn.commit()

    df_insert = df.copy()
    for col in df_insert.select_dtypes(include='object').columns:
        df_insert[col] = df_insert[col].apply(
            lambda x: str(x).encode('utf-8', 'ignore').decode('utf-8') if isinstance(x, str) else x
        )
    df_insert.to_sql('stock_pen', con=engine, if_exists='append', index=False)
    return {'inserted': len(df_insert), 'deleted': deleted}


# ── INSERT NEW ORDERS (simple template) ───────────────────────────────────────

# Cột bắt buộc
_REQUIRED_COLS = ['Kênh', 'Mã hàng', 'Tháng đặt', 'Năm đặt', 'SL đặt']
# Cột tùy chọn — tự điền nếu sản phẩm chưa có trong product_template
_OPTIONAL_META  = ['Danh mục', 'Danh mục con', 'Mã SP cha', 'Size']
# Cột tùy chọn — lịch sử trả hàng (mặc định 0)
_OPTIONAL_DEL   = ['SL TRẢ Năm Cũ'] + [f'SL TRẢ T{str(m).zfill(2)}' for m in range(1, 13)]
_CHANNEL_MAP    = {'CỬA HÀNG': 'KDC', 'BÁN SỈ': 'KDS'}


def get_order_template_bytes() -> bytes:
    """Trả về bytes của file Excel template với đầy đủ cột, có ghi chú hướng dẫn."""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Đơn đặt hàng'

    all_cols = _REQUIRED_COLS + _OPTIONAL_META + _OPTIONAL_DEL
    ws.append(all_cols)

    # Style header
    yellow = PatternFill('solid', fgColor='FFF2CC')   # bắt buộc
    green  = PatternFill('solid', fgColor='E2EFDA')   # meta tùy chọn
    blue   = PatternFill('solid', fgColor='DDEEFF')   # delivered tùy chọn
    bold   = Font(bold=True)
    for i, col in enumerate(all_cols, 1):
        cell = ws.cell(1, i)
        cell.font  = bold
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if col in _REQUIRED_COLS:
            cell.fill = yellow
        elif col in _OPTIONAL_META:
            cell.fill = green
        else:
            cell.fill = blue
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Ghi chú row 2
    ws.append(['# Vàng=bắt buộc', '# Xanh lá=tự điền nếu SP mới chưa có trong hệ thống',
               '', '', '', '', '', '', '', '# Xanh dương=lịch sử trả hàng, để trống = 0'])

    # Ví dụ SP đã có trong hệ thống (để trống meta → tự lookup)
    ws.append(['KDC', 'M4SUK7071', 6, 2026, 100, '', '', '', ''])
    # Ví dụ SP mới chưa có trong hệ thống (cần điền meta)
    ws.append(['ECOM', 'NEWPRD001', 6, 2026, 50, 'SNEAKERS', 'AIR MAX', 'NEWPRD', '39'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_order_template(uploaded_file, engine) -> tuple:
    """
    Parse Excel template đơn đặt hàng.
    - Cột bắt buộc: Kênh, Mã hàng, Tháng đặt, Năm đặt, SL đặt
    - Cột meta tùy chọn: Danh mục, Danh mục con, Mã SP cha, Size
      → ưu tiên dùng product_template; chỉ dùng giá trị Excel khi SP chưa có trong hệ thống
    - Cột delivered tùy chọn: SL TRẢ Năm Cũ, SL TRẢ T01..T12 (mặc định 0)
    Returns (df_ready, not_found_list).
    """
    df = pd.read_excel(uploaded_file)

    # Normalize tên cột
    col_aliases = {
        'kenh': 'Kênh', 'channel': 'Kênh',
        'ma hang': 'Mã hàng', 'fdcode': 'Mã hàng', 'ma_hang': 'Mã hàng',
        'thang dat': 'Tháng đặt', 'month': 'Tháng đặt', 'thang_dat': 'Tháng đặt',
        'nam dat': 'Năm đặt', 'year': 'Năm đặt', 'nam_dat': 'Năm đặt',
        'sl dat': 'SL đặt', 'qty': 'SL đặt', 'sl_dat': 'SL đặt',
        'danh muc': 'Danh mục', 'category': 'Danh mục',
        'danh muc con': 'Danh mục con', 'subcategory': 'Danh mục con',
        'ma sp cha': 'Mã SP cha', 'default_code': 'Mã SP cha',
        'size': 'Size',
        'sl tra nam cu': 'SL TRẢ Năm Cũ', 'delivered_old_year': 'SL TRẢ Năm Cũ',
        **{f'sl tra t{str(m).zfill(2)}': f'SL TRẢ T{str(m).zfill(2)}' for m in range(1, 13)},
    }
    df.columns = [col_aliases.get(c.lower().strip().replace('  ', ' '), c) for c in df.columns]

    # Bỏ dòng ghi chú (bắt đầu bằng #)
    if 'Kênh' in df.columns:
        df = df[~df['Kênh'].astype(str).str.startswith('#')].copy()

    missing = [c for c in _REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"File thiếu cột bắt buộc: {missing}")

    df = df.dropna(subset=['Mã hàng', 'Kênh']).copy()
    df['Kênh']      = df['Kênh'].astype(str).str.strip().replace(_CHANNEL_MAP)
    df['Mã hàng']   = df['Mã hàng'].astype(str).str.strip()
    df['Tháng đặt'] = df['Tháng đặt'].astype(int)
    df['Năm đặt']   = df['Năm đặt'].astype(int)
    df['SL đặt']    = df['SL đặt'].fillna(0).astype(int)

    valid_channels = {'KDC', 'KDS', 'ECOM'}
    bad = df[~df['Kênh'].isin(valid_channels)]
    if not bad.empty:
        raise ValueError(f"Kênh không hợp lệ: {bad['Kênh'].unique().tolist()}. Chỉ nhận: KDC, KDS, ECOM")

    # Enrich từ product_template — ưu tiên DB, fallback về giá trị Excel
    df_prod = (
        get_product_template(engine)[['fdcode', 'default_code', 'category', 'subcategory', 'size']]
        .drop_duplicates('fdcode')
        .set_index('fdcode')
    )

    def _get(row, prod_meta, db_key, excel_col):
        """Lấy từ product_template trước; nếu không có dùng giá trị Excel."""
        v = prod_meta.get(db_key)
        if v:
            return v
        if excel_col in df.columns:
            excel_val = row.get(excel_col)
            return str(excel_val).strip() if pd.notna(excel_val) and str(excel_val).strip() else None
        return None

    result_rows = []
    not_found = []
    for _, row in df.iterrows():
        fdc  = row['Mã hàng']
        meta = df_prod.loc[fdc].to_dict() if fdc in df_prod.index else {}
        if not meta:
            not_found.append(fdc)

        # Delivered history (optional cols, default 0)
        def _del(col):
            return int(row[col]) if col in df.columns and pd.notna(row.get(col)) else 0

        result_rows.append({
            'channel':               row['Kênh'],
            'fdcode':                fdc,
            'month_ord':             row['Tháng đặt'],
            'year_ord':              row['Năm đặt'],
            'qty_ord':               row['SL đặt'],
            'category':              _get(row, meta, 'category',     'Danh mục'),
            'subcategory':           _get(row, meta, 'subcategory',  'Danh mục con'),
            'default_code':          _get(row, meta, 'default_code', 'Mã SP cha'),
            'size':                  _get(row, meta, 'size',         'Size'),
            'delivered_old_year':    _del('SL TRẢ Năm Cũ'),
            **{f'delivered_{m}': _del(f'SL TRẢ T{str(m).zfill(2)}') for m in range(1, 13)},
            'qty_delivered_by_manu': 0,
            'order_pen':             row['SL đặt'],
        })

    df_out = pd.DataFrame(result_rows)
    return df_out, not_found


def get_existing_order_months(engine) -> pd.DataFrame:
    """Trả về các (year_ord, month_ord, channel, count) hiện có trong stock_pen."""
    delivered_sum = ' + '.join(
        ['delivered_old_year'] + [f'delivered_{m}' for m in range(1, 13)]
    )
    with engine.connect() as conn:
        return pd.read_sql(
            text(f"""
                SELECT year_ord, month_ord, channel,
                       COUNT(*) as so_dong,
                       SUM(qty_ord) as tong_sl_dat,
                       SUM({delivered_sum}) as tong_da_tra
                FROM stock_pen
                WHERE channel != 'NGOÀI ĐƠN'
                GROUP BY year_ord, month_ord, channel
                ORDER BY year_ord, month_ord, channel
            """),
            conn
        )


def get_orders_detail(engine, year: int, month: int, channel: str) -> pd.DataFrame:
    """Trả về chi tiết từng dòng đơn hàng của (year, month, channel), kèm cột has_delivery."""
    delivered_sum = ' + '.join(
        ['delivered_old_year'] + [f'delivered_{m}' for m in range(1, 13)]
    )
    with engine.connect() as conn:
        return pd.read_sql(text(f"""
            SELECT id, fdcode, default_code, size, qty_ord,
                   ({delivered_sum}) AS da_tra,
                   order_pen
            FROM stock_pen
            WHERE year_ord = :y AND month_ord = :m AND channel = :ch
              AND channel != 'NGOÀI ĐƠN'
            ORDER BY fdcode
        """), conn, params={'y': year, 'm': month, 'ch': channel})


def delete_orders_by_ids(engine, ids: list) -> dict:
    """Xóa các dòng theo danh sách id. Chỉ xóa dòng chưa có trả hàng."""
    if not ids:
        return {'deleted': 0, 'skipped': 0}
    delivered_sum = ' + '.join(
        ['delivered_old_year'] + [f'delivered_{m}' for m in range(1, 13)]
    )
    with engine.connect() as conn:
        # Lọc chỉ id chưa có trả hàng
        placeholders = ','.join([str(i) for i in ids])
        rows = conn.execute(text(f"""
            SELECT id FROM stock_pen
            WHERE id IN ({placeholders})
              AND ({delivered_sum}) = 0
        """)).fetchall()
        safe_ids = [r[0] for r in rows]
        skipped = len(ids) - len(safe_ids)

        deleted = 0
        if safe_ids:
            ph2 = ','.join([str(i) for i in safe_ids])
            r = conn.execute(text(f"DELETE FROM stock_pen WHERE id IN ({ph2})"))
            conn.commit()
            deleted = r.rowcount
    return {'deleted': deleted, 'skipped': skipped}


def delete_orders_by_month(engine, year: int, month: int, channel: str) -> dict:
    """
    Xóa đơn hàng theo (year_ord, month_ord, channel).
    Chỉ xóa khi chưa có trả hàng nào (tong_da_tra = 0).
    Returns dict: deleted, blocked (nếu đã có trả hàng).
    """
    delivered_sum = ' + '.join(
        ['delivered_old_year'] + [f'delivered_{m}' for m in range(1, 13)]
    )
    with engine.connect() as conn:
        # Kiểm tra xem có dòng nào đã được trả hàng không
        row = conn.execute(text(f"""
            SELECT COUNT(*) as total, SUM({delivered_sum}) as da_tra
            FROM stock_pen
            WHERE year_ord = :y AND month_ord = :m AND channel = :ch
              AND channel != 'NGOÀI ĐƠN'
        """), {'y': year, 'm': month, 'ch': channel}).fetchone()

        total   = int(row[0] or 0)
        da_tra  = int(row[1] or 0)

        if total == 0:
            return {'deleted': 0, 'blocked': False, 'message': 'Không tìm thấy đơn nào.'}

        if da_tra > 0:
            return {
                'deleted': 0, 'blocked': True,
                'message': f'Không thể xóa — đã có {da_tra:,} units được trả hàng. Dùng "Thay thế tháng" khi upload đơn mới để reset đúng cách.'
            }

        r = conn.execute(text("""
            DELETE FROM stock_pen
            WHERE year_ord = :y AND month_ord = :m AND channel = :ch
              AND channel != 'NGOÀI ĐƠN'
        """), {'y': year, 'm': month, 'ch': channel})
        conn.commit()
        return {'deleted': r.rowcount, 'blocked': False, 'message': f'Đã xóa {r.rowcount} dòng.'}
